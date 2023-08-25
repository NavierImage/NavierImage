import os 
import re
import openpyxl

wb = openpyxl.Workbook()
wb.create_sheet("only_oar")
wb.create_sheet("oar_num")
ws = wb["only_oar"]
ws2 = wb["oar_num"]


def get_contournames_info(contournames_path):
    with open(contournames_path, "r") as contournames:
        cnames_line_list = contournames.readlines()
    
    contour_dict = {}
    for idx, line in enumerate(cnames_line_list):
        line_str = line
        if "\n" in line_str:
            line_str = re.sub("\n", "", line_str)
        
        if line_str.isspace(): # space
            continue 
        elif line_str == "": # null
            continue
        
        number_det = re.findall(r"\d+", line_str) #Zn case except ex) Z8, Z1 ...
        check = 0
        if len(number_det) > 0:
            pass
        elif "gtv" in line_str.lower():
            check = 1
        else:
            if   line_str.lower() == "imported":
                pass
            elif line_str.lower() == "mrlcouch":
                pass
            elif line_str.lower() == "study ct":
                pass
            elif line_str.lower() == "mm":
                pass
            elif line_str.lower() == "patient":
                pass
            elif line_str.lower() == "general":
                pass
            elif line_str.lower() == "isocenter":
                pass
            else:
                check = 1
                
        
        if check == 1 and idx != len(cnames_line_list) -1:
            next_line = cnames_line_list[idx+1]
            contour_dict[line_str.lower()] = next_line.split(",")[0]
    return contour_dict

basepath = r"D:\!JUNG_newdata_T23D_mr\abdomenCT1"

pt_fol_list = os.listdir(basepath)

new_dict = {}
for i in range(len(pt_fol_list)):
    contournames_path = os.path.join(basepath, pt_fol_list[i] + r"/contournames")
    contournames = get_contournames_info(contournames_path)
    key_list = []
    
    for key, val in contournames.items():
        if "tv" in key.lower():
            key_list.append(key)
            try:
                new_dict[key] += 1
            except:
                new_dict[key] = 1
                
        if "prv" in key.lower():
            continue
        
        if "cochlea" in key.lower():
            print(pt_fol_list[i])
        if "femur" in key.lower():
            print(pt_fol_list[i])
        if "parotid" in key.lower():
            print(pt_fol_list[i])
        if "thyroid" in key.lower():
            print(pt_fol_list[i])
        
        
    key_list.sort()
    for j in range(len(key_list)):
        
        ws.cell(i+2, j+2, key_list[j]) 

row = 1
col = 1
for key, val in new_dict.items():
    ws2.cell(row, col, key)
    ws2.cell(row+1, col, val)
    col += 1
    
wb.save(r"D:\!JUNG_newdata_T23D_mr/contournames_gtv1.xlsx")